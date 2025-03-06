import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def add_batting_column(df, batting_value):
    df['batting'] = ''
    if not df.empty:
        df.at[0, 'batting'] = batting_value
    return df


def write_dataframes_to_excel(df1, df2_positive, df3_negative, df1_batting, df2_batting, df3_batting, fund_name,benchmark_name,
                              file_path, excess_return_data, final_scores):
    df1 = add_batting_column(df1, df1_batting)
    df2_positive = add_batting_column(df2_positive, df2_batting)
    df3_negative = add_batting_column(df3_negative, df3_batting)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        final_scores.to_excel(writer, sheet_name=f'{fund_name}_results', index=False)

        # Write individual DataFrames to separate sections of the sheet template
        df1.to_excel(writer, sheet_name='batting_average', startrow=0, startcol=0, index=False)
        df2_positive.to_excel(writer, sheet_name='batting_average', startrow=0, startcol=df1.shape[1] + 1, index=False)
        df3_negative.to_excel(writer, sheet_name='batting_average', startrow=0,
                              startcol=df1.shape[1] + df2_positive.shape[1] + 2, index=False)

        excess_return_data.to_excel(writer, sheet_name='excess', index=False)

    # Add title to the "results" sheet
    add_title_to_results(file_path, fund_name, benchmark_name)

    resize_columns(file_path, fund_name + '_results')
    resize_columns(file_path, 'batting_average')
    resize_columns(file_path, 'excess')
    add_borders_to_tables(file_path, fund_name + '_results', final_scores)
    add_borders_to_batting(file_path, 'batting_average', [df1, df2_positive, df3_negative])
    format_date_columns(file_path, 'batting_average')
    format_date_columns(file_path, 'excess')
    add_titles_to_batting(file_path, 'batting_average',
                          ['All Time Batting', 'Positive Benchmark', 'Negative Benchmark'],
                          [df1, df2_positive, df3_negative])
    insert_row_and_label_columns(file_path, 'batting_average',
                                 ['Date', 'Fund Return', 'Benchmark Return', 'Excess', 'Is Greater', 'Batting'])
    format_excess_table(file_path, 'excess')
    add_title_to_excess(file_path, 'excess', 'Excess Return Table')

def add_title_to_results(file_path, fund_name, benchmark_name):
    wb = load_workbook(file_path)
    ws = wb[f'{fund_name}_results']

    # Define the title formatting
    font = Font(color='FFFFFFFF', bold=True, size=16)  # White, bold, and large
    fill = PatternFill(start_color='FF10045A', end_color='FF10045A', fill_type='solid')  # Dark background
    alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells to create a wide enough area for the title
    last_column = ws.max_column  # Get the last column of data
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)

    # Add the title text
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{fund_name} vs {benchmark_name} results"
    title_cell.font = font
    title_cell.fill = fill
    title_cell.alignment = alignment

    # Optional: Adjust row height to make the title stand out more
    ws.row_dimensions[1].height = 30  # Adjust as needed

    wb.save(file_path)

def add_borders_to_tables(file_path, sheet_name, final_scores):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    num_rows_final_scores, num_cols_final_scores = final_scores.shape

    for row in ws.iter_rows(min_row=1, max_row=num_rows_final_scores + 1, min_col=1, max_col=num_cols_final_scores):
        for cell in row:
            cell.border = border

    wb.save(file_path)


def add_borders_to_batting(file_path, sheet_name, dfs):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    start_col = 1
    for df in dfs:
        num_rows, num_cols = df.shape
        for row in ws.iter_rows(min_row=4, max_row=num_rows + 4, min_col=start_col, max_col=start_col + num_cols - 1):
            for cell in row:
                cell.border = border
        start_col += num_cols + 2  # Ensure there's a full-width empty column between tables

    wb.save(file_path)


def resize_columns(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Get all merged cell ranges in the worksheet
    merged_ranges = ws.merged_cells.ranges

    # Loop through all columns
    for col in ws.columns:
        if col:  # Check if the column is not empty
            skip_column = False

            # Check if this column is part of any merged range
            for merged_range in merged_ranges:
                # If the column is part of the merged range, skip it
                if col[0].row >= merged_range.min_row and col[0].row <= merged_range.max_row:
                    if col[0].column >= merged_range.min_col and col[0].column <= merged_range.max_col:
                        skip_column = True
                        break

            if not skip_column:  # Only proceed if the column is not part of a merged range
                max_length = 0
                column = col[0].column_letter  # Get the column letter of the first cell in the column
                for cell in col:
                    try:
                        if isinstance(cell.value, (str, int, float)):
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        if cell.is_date:
                            max_length = max(max_length, 10)  # Date format length
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Adjusting width with some padding
                ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)




def format_date_columns(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for col in ws.columns:
        for cell in col:
            if cell.is_date:
                cell.number_format = 'yyyy-mm-dd'

    wb.save(file_path)


def add_titles_to_batting(file_path, sheet_name, titles, dfs):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    font = Font(color='FFFFFFFF', bold=True, size=16)  # ARGB format
    fill = PatternFill(start_color='FF10045A', end_color='FF10045A', fill_type='solid')  # ARGB format
    alignment = Alignment(horizontal='center', vertical='center')
    black_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')  # ARGB format

    start_col = 1
    for title, df in zip(titles, dfs):
        num_cols = df.shape[1]
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + num_cols - 1)
        cell = ws.cell(row=1, column=start_col)
        cell.value = title
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment

        start_col += num_cols + 1
        ws.column_dimensions[get_column_letter(start_col - 1)].width = 1
        ws.merge_cells(start_row=1, start_column=start_col - 1, end_row=4, end_column=start_col - 1)
        for row in range(1, 5):  # Fill all four rows of the black column
            cell = ws.cell(row=row, column=start_col - 1)
            cell.fill = black_fill

    wb.save(file_path)


def insert_row_and_label_columns(file_path, sheet_name, columns):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    ws.insert_rows(2)  # Insert a row for the headers

    header_fill = PatternFill(start_color='FF443793', end_color='FF443793', fill_type='solid')  # ARGB format
    header_font = Font(color='FFFFFFFF')  # ARGB format
    data_font = Font(color='ff10045a')  # ARGB format

    start_col = 1
    for col_group in range(3):  # Assuming three tables
        for col_num, column in enumerate(columns, start=start_col):
            cell = ws.cell(row=2, column=col_num)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=start_col, max_col=start_col + len(columns) - 1):
            for cell in row:
                cell.font = data_font
        start_col += len(columns) + 1  # Adjust the starting column for the next table

    wb.save(file_path)


def format_excess_table(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Define header style to match the batting_average format
    header_fill = PatternFill(start_color='FF443793', end_color='FF443793', fill_type='solid')  # ARGB format
    header_font = Font(color='FFFFFFFF', bold=True)  # ARGB format, bold header font
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Define border style
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Format the header row (row 2) - Assuming headers are in row 2
    for col_num, cell in enumerate(ws[2], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Resize columns based on content (same as before)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if isinstance(cell.value, (str, int, float)):
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                if cell.is_date:
                    max_length = max(max_length, 10)  # Date format length
            except:
                pass
        adjusted_width = (max_length + 2)  # Adjusting width with some padding
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)


def add_title_to_excess(file_path, sheet_name, title):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Define the title cell formatting (same as batting average titles)
    font = Font(color='FFFFFFFF', bold=True, size=16)  # ARGB format for white and bold text
    fill = PatternFill(start_color='FF10045A', end_color='FF10045A', fill_type='solid')  # ARGB format for fill color
    alignment = Alignment(horizontal='center', vertical='center')

    # Shift the entire table down by one row (starting from row 2 onwards)
    for row in range(ws.max_row, 0, -1):  # Start from the last row and go upwards
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row + 1, column=col).value = ws.cell(row=row, column=col).value

    # Merge cells in row 1 for the title
    last_column = ws.max_column  # Get the last column with data
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)

    # Set the title text, font, fill, and alignment
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = font
    title_cell.fill = fill
    title_cell.alignment = alignment

    # Optional: Set the row height for the title to make it stand out more
    ws.row_dimensions[1].height = 30  # Adjust this height to make the title more prominent

    wb.save(file_path)

    wb.save(file_path)
